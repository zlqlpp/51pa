######excel 2003   2007版本的  读和 写功能
# -*- coding: utf-8 -*-
#https://blog.csdn.net/jeikerxiao/article/details/73614335
# 读写2003 excel
import xlrd
import xlwt
# 读写2007 excel
import openpyxl


def write03Excel(path):
    wb = xlwt.Workbook()
    sheet = wb.add_sheet("2003测试表")
    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(path)
    print("写入数据成功！")


def read03Excel(path):
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    for i in range(0, worksheet.nrows):
        row = worksheet.row(i)
        #for j in range(0, worksheet.ncols):
        #    print(worksheet.cell_value(i, j), "\t", end="")
        #print()
        print(worksheet.cell_value(i, 0)+worksheet.cell_value(i, 1)+worksheet.cell_value(i, 2)+worksheet.cell_value(i, 3)+worksheet.cell_value(i, 4))


def write07Excel(path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '2007测试表'

    value = [["名称", "价格", "出版社", "语言"],
             ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
             ["暗时间", "32.4", "人民邮电出版社", "中文"],
             ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, 4):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))

    wb.save(path)
    print("写入数据成功！")


def read07Excel(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('2007测试表')

    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


file_2007 = 'data/2007.xlsx'

#write03Excel(file_2003)
read03Excel('/root/lllllffffff.xls')

#write07Excel(file_2007)
#read07Excel(file_2007)
